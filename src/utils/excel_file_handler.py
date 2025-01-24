import pandas
from openpyxl import load_workbook
import sys
import os

base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # Lấy thư mục gốc của dự án
excel_path = os.path.join(base_dir, "ClassManager.xlsx")

class ExcelFileHandler:
    @staticmethod
    def read_excel(sheet_name=None):
        # Đọc file Excel có sẵn  
        try:
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"File không tồn tại: {excel_path}")
            
            if sheet_name is not None:
                return pandas.read_excel(excel_path, sheet_name)
            return [pandas.read_excel(excel_path, sheet_name) for sheet_name in pandas.ExcelFile(excel_path).sheet_names]
        
        except FileNotFoundError as e:
            print(e)
            return None
      
    @staticmethod
    def save_excel(df):
        """
        Lưu dữ liệu vào file Excel.
        """
        try:
            # Ghi dữ liệu vào file Excel
            df.to_excel(excel_path, index=False)
            
            # Mở và lưu lại file bằng openpyxl để tránh lỗi định dạng
            wb = load_workbook(excel_path)
            wb.save(excel_path)
            print(f"Lưu file thành công: {excel_path}")
        except Exception as e:
            print(f"Lỗi khi lưu file Excel: {e}")
            
    @staticmethod
    def add_sheet(sheet_name):
        """
        Tạo thêm một sheet mới trong file Excel.
        """
        try:
            wb = load_workbook(excel_path)
            if sheet_name in wb.sheetnames:
                print(f"Sheet '{sheet_name}' đã tồn tại.")
                return

            wb.create_sheet(sheet_name)
            wb.save(excel_path)
            print(f"Đã thêm sheet mới: {sheet_name}")
        except Exception as e:
            print(f"Lỗi khi thêm sheet: {e}")
            
    @staticmethod
    def delete_sheet(sheet_name):
        """
        Xóa một sheet trong file Excel.
        """
        try:
            wb = load_workbook(excel_path)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' không tồn tại.")
            
            # Không cho phép xóa nếu chỉ còn một sheet
            if len(wb.sheetnames) == 1:
                raise ValueError("Không thể xóa sheet cuối cùng trong file.")
            
            # Xóa sheet
            del wb[sheet_name]
            wb.save(excel_path)
            print(f"Đã xóa sheet '{sheet_name}'.")
        except Exception as e:
            print(f"Lỗi khi xóa sheet: {e}")
            
    @staticmethod
    def write_cell(sheet_name, cell, value):
        """
        Ghi dữ liệu vào một ô chỉ định trong sheet.
        """
        try:
            wb = load_workbook(excel_path)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' không tồn tại.")

            sheet = wb[sheet_name]
            sheet[cell] = value
            wb.save(excel_path)
            print(f"Đã ghi dữ liệu '{value}' vào ô {cell} trong sheet '{sheet_name}'.")
        except Exception as e:
            print(f"Lỗi khi ghi dữ liệu vào ô: {e}")

    @staticmethod
    def delete_cell(sheet_name, cell):
        """
        Xóa dữ liệu của một ô chỉ định trong sheet.
        """
        try:
            wb = load_workbook(excel_path)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' không tồn tại.")

            sheet = wb[sheet_name]
            sheet[cell] = None
            wb.save(excel_path)
            print(f"Đã xóa dữ liệu tại ô {cell} trong sheet '{sheet_name}'.")
        except Exception as e:
            print(f"Lỗi khi xóa dữ liệu tại ô: {e}")
            
    @staticmethod
    def delete_row(sheet_name, row_number):
        """
        Xóa một hàng trong sheet chỉ định.
        """
        try:
            wb = load_workbook(excel_path)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' không tồn tại.")

            sheet = wb[sheet_name]
            max_row = sheet.max_row
            if row_number < 1 or row_number > max_row:
                raise ValueError(f"Hàng {row_number} không hợp lệ. Tổng số hàng: {max_row}.")

            sheet.delete_rows(row_number)
            wb.save(excel_path)
            print(f"Đã xóa hàng {row_number} trong sheet '{sheet_name}'.")
        except Exception as e:
            print(f"Lỗi khi xóa hàng: {e}")

    @staticmethod
    def delete_column(sheet_name, column_number):
        """
        Xóa một cột trong sheet chỉ định.
        """
        try:
            wb = load_workbook(excel_path)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' không tồn tại.")

            sheet = wb[sheet_name]
            max_column = sheet.max_column
            if column_number < 1 or column_number > max_column:
                raise ValueError(f"Cột {column_number} không hợp lệ. Tổng số cột: {max_column}.")

            sheet.delete_cols(column_number)
            wb.save(excel_path)
            print(f"Đã xóa cột {column_number} trong sheet '{sheet_name}'.")
        except Exception as e:
            print(f"Lỗi khi xóa cột: {e}")
            
